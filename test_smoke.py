import tempfile
import uuid
import re
import json
from io import BytesIO
from pathlib import Path
from zipfile import ZipFile

import httpx
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from app import FEISHU_AGENT_LOGS_DIR, ProjectPayload, app, call_chat_completion, extract_ai_response_content, extract_document_text, handle_feishu_message, load_config, load_users, market_record_path, merge_register_rows, parse_json_text, project_path, resolve_source_excerpt, save_config, save_project, save_users, ws_message_to_payload


def cleanup(project_id: str) -> None:
    project_path(project_id).unlink(missing_ok=True)


def cleanup_user(username: str) -> None:
    payload = load_users()
    payload['users'] = [user for user in payload.get('users', []) if user.get('username') != username]
    save_users(payload)


def cleanup_market(kind: str, record_id: str) -> None:
    market_record_path(kind, record_id).unlink(missing_ok=True)


sample = parse_json_text('```json {"ok": true} ```')
assert sample['ok'] is True
assert extract_ai_response_content({'choices': [{'message': {'content': '{"ok":true}'}}]}) == '{"ok":true}'
assert extract_ai_response_content({'choices': [{'message': {'content': [{'type': 'text', 'text': '{"ok":true}'}]}}]}) == '{"ok":true}'
assert extract_ai_response_content({'choices': [{'message': {'tool_calls': [{'function': {'arguments': '{"ok":true}'}}]}}]}) == '{"ok":true}'
assert extract_ai_response_content({'output': [{'content': [{'type': 'output_text', 'text': '{"ok":true}'}]}]}) == '{"ok":true}'
assert extract_ai_response_content({'data': {'result': '{"ok":true}'}}) == '{"ok":true}'

merged = merge_register_rows(
    [
        {'project_name': '项目A', 'bid_no': '01', 'deposit_amount': '1000', 'remark': ''},
        {'project_name': '项目A', 'bid_no': '02', 'deposit_amount': '2000', 'remark': ''},
    ]
)
assert '02' in merged['bid_no']
assert '2 个标段' in merged['remark']
assert resolve_source_excerpt('股份有限公司', '...', '福建省东南电化股份有限公司') == '福建省东南电化股份有限公司'

with tempfile.TemporaryDirectory() as temp_dir:
    xlsx_path = Path(temp_dir) / 'sample.xlsx'
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = '招标信息'
    sheet['A1'] = '项目名称'
    sheet['B1'] = '测试招标项目'
    sheet['A2'] = '投标截止'
    sheet['B2'] = '2026-03-04 09:00'
    workbook.save(xlsx_path)
    workbook.close()
    extracted = extract_document_text(xlsx_path, '.xlsx')
    assert '工作表：招标信息' in extracted
    assert '项目名称 | 测试招标项目' in extracted


class DummyReadTimeoutClient:
    calls = 0

    @staticmethod
    def post(*args, **kwargs):
        DummyReadTimeoutClient.calls += 1
        raise httpx.ReadTimeout('simulated read timeout')


original_httpx_post = httpx.post
httpx.post = DummyReadTimeoutClient.post
try:
    try:
        call_chat_completion(
            {'base_url': 'https://example.com/v1', 'api_key': 'sk-test', 'model': 'demo', 'temperature': 0.1},
            system_prompt='test',
            user_prompt='test',
        )
        raise AssertionError('expected timeout')
    except Exception as exc:
        assert '响应超时' in str(exc.detail)
        assert DummyReadTimeoutClient.calls == 1
finally:
    httpx.post = original_httpx_post

project_id = 'f' * 32
cleanup(project_id)
save_project(
    project_id,
    ProjectPayload(
        title='测试项目',
        result={
            'document_summary': {'project_name': '测试项目'},
            'register_rows': [{'project_name': '测试项目'}],
            'document_text_full': '测试项目原文',
        },
    ),
)
assert project_path(project_id).exists()
cleanup(project_id)

client = TestClient(app)
cleanup_user('demo_user')

root_resp = client.get('/', follow_redirects=False)
assert root_resp.status_code in (302, 307)
assert root_resp.headers['location'] == '/login'

app_redirect_resp = client.get('/app', follow_redirects=False)
assert app_redirect_resp.status_code in (302, 307)
assert app_redirect_resp.headers['location'] == '/login'

login_resp = client.post('/api/auth/login', json={'username': 'ruico', 'password': 'Ruico668@'})
assert login_resp.status_code == 200
token = login_resp.json()['token']
auth_headers = {'Authorization': f'Bearer {token}'}
assert 'bid_parser_token=' in (login_resp.headers.get('set-cookie') or '')

app_resp = client.get('/app', headers=auth_headers)
assert app_resp.status_code == 200

project_payload = {
    'title': '接口测试项目',
    'source_file_name': 'demo.pdf',
    'register_mode': 'packages',
    'sheet_name': '2026',
    'follow_up': {
        'bid_status': '已投标',
        'award_status': '已中标',
        'award_date': '2026-03-10',
        'award_company': '我司',
        'our_award_amount': '1280000',
        'competitor_award_amount': '1310000',
        'information_source': '业务补录',
        'tracking_note': '完成中标补录',
        'register_year': '2026',
    },
    'our_quotes': [
        {
            'package_name': '合同包1',
            'round_no': 1,
            'quote_date': '2026-03-01',
            'quote_company': '我司',
            'currency': 'CNY',
            'tax_mode': '含税',
            'unit_price': '650',
            'total_price': '1280000',
            'is_submitted': True,
            'is_awarded': True,
            'remark': '最终中标价',
        }
    ],
    'competitor_quotes': [
        {
            'quote_company': '竞对A',
            'package_name': '合同包1',
            'quote_date': '2026-03-01',
            'currency': 'CNY',
            'unit_price': '665',
            'total_price': '1310000',
            'ranking': '2',
            'is_awarded': False,
            'source': '中标公示',
            'remark': '次低价',
        }
    ],
    'timeline': [
        {'date': '2026-01-27', 'type': 'parse', 'note': '完成标书解析'},
        {'date': '2026-03-10', 'type': 'award', 'note': '完成中标结果补录'},
    ],
    'result': {
        'document_summary': {
            'project_name': '福建省东南电化股份有限公司2026年液碱海路运输业务',
            'bid_no': 'WHTY/B-QT-2026-EB008-01',
            'tenderer': '福建省东南电化股份有限公司',
            'bid_deadline': '2026/3/4 9:00',
            'open_time': '2026/3/4 9:00',
            'submission_method': '投标平台递交',
            'deposit_amount': '10万',
            'service_period': '2026年3月-2027年3月',
            'qualification_requirements': '资质齐全',
            'vessel_requirements': '适装适靠',
            'technical_business': '按招标文件执行',
            'quotation': '固定单价',
            'evaluation_method': '综合评分法',
        },
        'extraction_fields': {
            'open_time': {'value': '2026/3/4 9:00', 'source_excerpt': '开标时间：2026/3/4 9:00'},
            'submission_method': {'value': '投标平台递交', 'source_excerpt': '投标文件递交方式：投标平台递交'},
        },
        'register_rows': [
            {
                'receive_date': '2026/1/27',
                'project_name': '福建省东南电化股份有限公司2026年液碱海路运输业务',
                'is_awarded': '√',
                'bid_no': 'WHTY/B-QT-2026-EB008-01',
                'tenderer': '福建省东南电化股份有限公司',
                'bid_deadline': '2026/3/4 9:00',
                'deposit_amount': '10万',
                'contract_note': '2026年3月-2027年3月',
                'deposit_return_status': '已退还10万元（履约保证金）',
                'payment_method': '对公转账',
                'acquisition_method': '万华电子招标投标交易网',
                'submission_method': '投标平台递交',
                'remark': '',
            }
        ],
        'analysis': {
            'summary': '测试摘要',
            'qualification_files': ['营业执照'],
            'business_points': ['注意报价口径'],
            'scoring_points': ['关注服务方案'],
            'risks': ['需人工确认船期'],
        },
        'match_review': [
            {
                'area': '业务匹配',
                'item': '运力核对',
                'status': '需人工确认',
                'reason': '未接内部运力系统',
                'action': '人工确认可用运力',
            }
        ],
        'document_text_excerpt': '示例原文',
        'document_text_full': '投标文件递交方式：投标平台递交\n开标时间：2026/3/4 9:00\n福建省东南电化股份有限公司',
    },
}

create_resp = client.post('/api/projects', json=project_payload, headers=auth_headers)
assert create_resp.status_code == 200
created = create_resp.json()
created_id = created['project_id']

list_resp = client.get('/api/projects?q=接口测试', headers=auth_headers)
assert list_resp.status_code == 200
assert any(item['project_id'] == created_id for item in list_resp.json()['items'])
assert list_resp.json()['stats']['awarded_projects'] >= 1

get_resp = client.get(f'/api/projects/{created_id}', headers=auth_headers)
assert get_resp.status_code == 200
project = get_resp.json()
result = project['result']

overview_resp = client.post(
    '/api/export/overview',
    json={
        'title': project['title'],
        'result': result,
        'follow_up': project['follow_up'],
        'our_quotes': project['our_quotes'],
        'competitor_quotes': project['competitor_quotes'],
        'timeline': project['timeline'],
    },
    headers=auth_headers,
)
assert overview_resp.status_code == 200
wb_overview = load_workbook(BytesIO(overview_resp.content))
assert wb_overview['解析总览']['A1'].value == '标书解析总览'
assert wb_overview['解析总览']['B2'].value == '接口测试项目'
assert wb_overview['解析总览']['A4'].value == '核心信息'

extraction_resp = client.post('/api/export/extraction', json={'result': result}, headers=auth_headers)
assert extraction_resp.status_code == 200
wb = load_workbook(BytesIO(extraction_resp.content))
assert wb.sheetnames[0] == '摘取结果'
assert '原文依据' in wb.sheetnames
assert wb['摘取结果']['B2'].value == '2026/3/4 9:00'
assert wb['原文依据']['B2'].value == '开标时间：2026/3/4 9:00'

legacy_extraction_resp = client.post('/api/export/extraction', json={'extraction_fields': result['extraction_fields']}, headers=auth_headers)
assert legacy_extraction_resp.status_code == 200

register_resp = client.post('/api/export/register', json={'rows': result['register_rows'], 'sheet_name': '2028'}, headers=auth_headers)
assert register_resp.status_code == 200
wb2 = load_workbook(BytesIO(register_resp.content))
assert '2028' in wb2.sheetnames
sheet = wb2['2028']
assert sheet['C3'].value
assert sheet['B3'].number_format == 'yyyy/m/d'
assert sheet['G3'].number_format == 'yyyy/m/d h:mm'
assert sheet['H3'].value == 100000
assert sheet['A3'].border.left.style == 'thin'
assert sheet['N3'].alignment.horizontal == 'center'

our_quotes_resp = client.post(
    '/api/export/our-quotes',
    json={'title': project['title'], 'follow_up': project['follow_up'], 'rows': project['our_quotes']},
    headers=auth_headers,
)
assert our_quotes_resp.status_code == 200
wb3 = load_workbook(BytesIO(our_quotes_resp.content))
assert wb3['报价一览']['A1'].value == '我司报价一览表'
assert wb3['报价一览']['E5'].value == '我司'

competitor_quotes_resp = client.post(
    '/api/export/competitor-quotes',
    json={'title': project['title'], 'follow_up': project['follow_up'], 'rows': project['competitor_quotes']},
    headers=auth_headers,
)
assert competitor_quotes_resp.status_code == 200
wb4 = load_workbook(BytesIO(competitor_quotes_resp.content))
assert wb4['报价一览']['A1'].value == '竞对报价对比'
assert wb4['报价一览']['B5'].value == '竞对A'

ledger_resp = client.post(
    '/api/export/ledger',
    json={'q': '接口测试', 'year': '2026', 'bid_status': '已投标', 'award_status': '已中标'},
    headers=auth_headers,
)
assert ledger_resp.status_code == 200
wb5 = load_workbook(BytesIO(ledger_resp.content))
assert wb5['历史台账']['A1'].value == '历史台账汇总'
assert wb5['历史台账']['D2'].value == '2026'
assert wb5['历史台账']['F2'].value == '已投标'
assert wb5['历史台账']['H2'].value == '已中标'
assert wb5['历史台账']['B6'].value == '接口测试项目'

cargo_a = {
    'board_type': '\u5373\u65f6\u8d27\u76d8',
    'segment': '\u5185\u8d38\u5316',
    'cargo_name': 'smoke-benzene-market-filter',
    'tonnage': '4000\u5428',
    'load_port': '\u5f20\u5bb6\u6e2f',
    'discharge_port': '\u4e1c\u839e',
    'cargo_owner': '\u8fdc\u5927',
    'status': '\u8ddf\u8fdb\u4e2d',
    'raw_text': 'raw text mentions smoke-methane-market-filter only',
}
cargo_b = {**cargo_a, 'cargo_name': 'smoke-methane-market-filter', 'raw_text': 'same as structured'}
cargo_a_resp = client.post('/api/market-skill/cargo', json={'kind': 'cargo', 'record': cargo_a}, headers=auth_headers)
cargo_b_resp = client.post('/api/market-skill/cargo', json={'kind': 'cargo', 'record': cargo_b}, headers=auth_headers)
assert cargo_a_resp.status_code == 200
assert cargo_b_resp.status_code == 200
cargo_a_id = cargo_a_resp.json()['record']['id']
cargo_b_id = cargo_b_resp.json()['record']['id']
try:
    methane_resp = client.get(
        '/api/market-skill/cargo',
        params={'q': 'smoke-methane-market-filter', 'segment': '\u5185\u8d38\u5316', 'status': '\u8ddf\u8fdb\u4e2d'},
        headers=auth_headers,
    )
    benzene_resp = client.get(
        '/api/market-skill/cargo',
        params={'q': 'smoke-benzene-market-filter', 'segment': '\u5185\u8d38\u5316', 'status': '\u8ddf\u8fdb\u4e2d'},
        headers=auth_headers,
    )
    assert methane_resp.status_code == 200
    assert benzene_resp.status_code == 200
    assert [item['id'] for item in methane_resp.json()['items']] == [cargo_b_id]
    assert [item['id'] for item in benzene_resp.json()['items']] == [cargo_a_id]
finally:
    cleanup_market('cargo', cargo_a_id)
    cleanup_market('cargo', cargo_b_id)

cargo_deal = {
    'board_type': '\u5373\u65f6\u8d27\u76d8',
    'segment': '\u5185\u8d38\u5316',
    'cargo_name': 'smoke-trend-xylene',
    'cargo_standard_name': 'smoke-trend-xylene',
    'tonnage': '3000\u5428',
    'load_port': '\u5b81\u6ce2',
    'discharge_port': '\u4e1c\u839e',
    'route': '\u5b81\u6ce2 - \u4e1c\u839e',
    'cargo_date': '2026-06-01',
    'deal_date': '2026-06-15',
    'status': '\u5df2\u6210\u4ea4',
    'deal_price': '128',
    'executing_vessel': 'smoke-vessel-008',
    'price_unit': '\u5143/\u5428',
    'currency': 'CNY',
}
cargo_open = {**cargo_deal, 'status': '\u8ddf\u8fdb\u4e2d', 'deal_price': '999', 'cargo_name': 'smoke-trend-open'}
cargo_deal_resp = client.post('/api/market-skill/cargo', json={'kind': 'cargo', 'record': cargo_deal}, headers=auth_headers)
cargo_open_resp = client.post('/api/market-skill/cargo', json={'kind': 'cargo', 'record': cargo_open}, headers=auth_headers)
assert cargo_deal_resp.status_code == 200
assert cargo_open_resp.status_code == 200
cargo_deal_id = cargo_deal_resp.json()['record']['id']
cargo_open_id = cargo_open_resp.json()['record']['id']
try:
    cargo_report_resp = client.post(
        '/api/market-skill/report',
        json={
            'kind': 'cargo',
            'period': 'monthly',
            'start_date': '2026-06-01',
            'end_date': '2026-06-30',
            'filters': {'q': 'smoke-trend'},
        },
        headers=auth_headers,
    )
    assert cargo_report_resp.status_code == 200
    cargo_report = cargo_report_resp.json()
    assert cargo_report['source_stats']['deal_count'] == 1
    assert cargo_report['trend_points'][0]['avg_price'] == 128
    assert cargo_report['detail_rows'][0]['price'] == 128
    assert cargo_report['detail_rows'][0]['executing_vessel'] == 'smoke-vessel-008'
    cargo_report_export_resp = client.post(
        '/api/export/market-skill-report',
        json={'kind': 'cargo', 'report': cargo_report, 'format': 'xlsx'},
        headers=auth_headers,
    )
    assert cargo_report_export_resp.status_code == 200
    wb6 = load_workbook(BytesIO(cargo_report_export_resp.content))
    assert wb6['分析报告']['A1'].value == '商机市场分析报告'
    cargo_docx_resp = client.post(
        '/api/export/market-skill-report',
        json={'kind': 'cargo', 'report': cargo_report},
        headers=auth_headers,
    )
    assert cargo_docx_resp.status_code == 200
    with ZipFile(BytesIO(cargo_docx_resp.content)) as docx:
        names = set(docx.namelist())
        assert {'[Content_Types].xml', '_rels/.rels', 'word/document.xml', 'word/styles.xml'} <= names
        assert '商机市场 AI 分析报告' in docx.read('word/document.xml').decode('utf-8')
finally:
    cleanup_market('cargo', cargo_deal_id)
    cleanup_market('cargo', cargo_open_id)

import app as app_module_for_market_update
old_market_extract_json = app_module_for_market_update.market_extract_json
app_module_for_market_update.market_extract_json = lambda kind, text: None
old_cargo_id = ''
try:
    old_cargo_resp = client.post(
        '/api/market-skill/cargo',
        json={'kind': 'cargo', 'record': {
            'cargo_name': 'smoke-existing-cargo',
            'tonnage': '3000吨',
            'load_port': '宁波',
            'discharge_port': '珠海',
            'cargo_owner': '旧货主',
            'status': '跟进中',
            'executing_vessel': '旧船001',
        }},
        headers=auth_headers,
    )
    assert old_cargo_resp.status_code == 200
    old_cargo = old_cargo_resp.json()['record']
    old_cargo_id = old_cargo['id']
    extract_update_resp = client.post(
        '/api/market-skill/extract',
        json={
            'kind': 'cargo',
            'record_id': old_cargo_id,
            'current_record': old_cargo,
            'text': '最终确定使用恒州008，最终成交价175',
        },
        headers=auth_headers,
    )
    assert extract_update_resp.status_code == 200
    extracted_update = extract_update_resp.json()
    assert extracted_update['mode'] == 'update'
    assert extracted_update['record']['id'] == old_cargo_id
    assert extracted_update['record']['cargo_name'] == 'smoke-existing-cargo'
    assert extracted_update['record']['load_port'] == '宁波'
    assert extracted_update['record']['executing_vessel'] == '恒州008'
    assert extracted_update['record']['deal_price'] == '175'
    records_before_update = client.get('/api/market-skill/cargo', params={'q': 'smoke-existing-cargo'}, headers=auth_headers).json()['items']
    assert [item['id'] for item in records_before_update] == [old_cargo_id]
    update_save_resp = client.put(
        f'/api/market-skill/cargo/{old_cargo_id}',
        json={'kind': 'cargo', 'record': extracted_update['record']},
        headers=auth_headers,
    )
    assert update_save_resp.status_code == 200
    records_after_update = client.get('/api/market-skill/cargo', params={'q': 'smoke-existing-cargo'}, headers=auth_headers).json()['items']
    assert [item['id'] for item in records_after_update] == [old_cargo_id]
    after_update = client.get(f'/api/market-skill/cargo/{old_cargo_id}', headers=auth_headers)
    assert after_update.status_code == 200
    version_before = after_update.json()['record']
    assert version_before['executing_vessel'] == '恒州008'
    concurrent_update = client.put(
        f'/api/market-skill/cargo/{old_cargo_id}',
        json={'kind': 'cargo', 'record': {**version_before, 'remark': '并发更新'}},
        headers=auth_headers,
    )
    assert concurrent_update.status_code == 200
    assert concurrent_update.json()['record']['updated_at'] != version_before['updated_at']
    stale_update = client.put(
        f'/api/market-skill/cargo/{old_cargo_id}',
        json={'kind': 'cargo', 'record': {**version_before, 'executing_vessel': '不应覆盖'}},
        headers=auth_headers,
    )
    assert stale_update.status_code == 409
finally:
    app_module_for_market_update.market_extract_json = old_market_extract_json
    if old_cargo_id:
        cleanup_market('cargo', old_cargo_id)

agent_original_config = load_config()
agent_original_call = app_module_for_market_update.call_chat_completion
agent_cargo_id = ''
agent_conversation_id = ''
try:
    save_config({**agent_original_config, 'base_url': 'https://example.com/v1', 'api_key': 'sk-agent-smoke', 'model': 'agent-smoke'})
    agent_record_resp = client.post(
        '/api/market-skill/cargo',
        json={'kind': 'cargo', 'record': {'cargo_name': 'smoke-agent-cargo', 'load_port': '宁波', 'discharge_port': '珠海', 'tonnage': '3000吨', 'status': '跟进中', 'executing_vessel': '旧船002'}},
        headers=auth_headers,
    )
    assert agent_record_resp.status_code == 200
    agent_record = agent_record_resp.json()['record']
    agent_cargo_id = agent_record['id']

    def fake_web_agent_chat(config, *, system_prompt, user_prompt, temperature=None):
        if 'available_tools' in user_prompt:
            return json.dumps({'reply': '', 'tool_calls': [{'name': 'draft_market_record_update', 'arguments': {'kind': 'cargo', 'record_id': agent_cargo_id, 'text': '最终执行船舶改成恒州008'}}], 'needs_confirmation': False, 'pending_update': {}}, ensure_ascii=False)
        if 'available_tools' not in user_prompt and '待确认' in user_prompt:
            return '已生成当前货盘的变更预览，请确认保存。'
        return '已保存当前货盘的变更。'

    app_module_for_market_update.market_extract_json = lambda kind, text: None
    app_module_for_market_update.call_chat_completion = fake_web_agent_chat
    page_context = json.dumps({'workspace': 'market', 'market_kind': 'cargo', 'current_market_record_id': agent_cargo_id, 'market_record': agent_record})
    agent_chat_resp = client.post(
        '/api/agent/chat',
        data={'message': '把当前货盘最终执行船舶改成恒州008', 'page_context': page_context},
        headers=auth_headers,
    )
    assert agent_chat_resp.status_code == 200
    agent_result = agent_chat_resp.json()
    agent_conversation_id = agent_result['conversation_id']
    assert agent_result['pending_change']['record_id'] == agent_cargo_id
    assert any(action['type'] == 'patch_market_form' for action in agent_result['actions'])
    assert '恒州008' in agent_result['answer']

    confirm_resp = client.post(
        '/api/agent/chat',
        data={'message': '确认保存', 'conversation_id': agent_conversation_id, 'page_context': page_context},
        headers=auth_headers,
    )
    assert confirm_resp.status_code == 200
    assert '保存' in confirm_resp.json()['answer']
    saved_agent_record = client.get(f'/api/market-skill/cargo/{agent_cargo_id}', headers=auth_headers).json()['record']
    assert saved_agent_record['executing_vessel'] == '恒州008'
    refused_resp = client.post(
        '/api/agent/chat',
        data={'message': '删除当前货盘', 'conversation_id': agent_conversation_id, 'page_context': page_context},
        headers=auth_headers,
    )
    assert refused_resp.status_code == 200
    assert '不能删除记录' in refused_resp.json()['answer']
    assert client.get(f'/api/market-skill/cargo/{agent_cargo_id}', headers=auth_headers).status_code == 200
finally:
    app_module_for_market_update.market_extract_json = old_market_extract_json
    app_module_for_market_update.call_chat_completion = agent_original_call
    save_config(agent_original_config)
    if agent_cargo_id:
        cleanup_market('cargo', agent_cargo_id)
    if agent_conversation_id:
        (app_module_for_market_update.AGENT_SESSIONS_DIR / f'{agent_conversation_id}.json').unlink(missing_ok=True)

newbuilding_record = {
    'stage': '\u5df2\u5b8c\u9020\u5e76\u51fa\u5382\u6295\u8fd0',
    'ship_name': 'smoke-newbuilding-1',
    'update_date': '2026-06-02',
    'status_update_date': '2026-06-20',
    'shipyard': '\u6c5f\u5357\u9020\u8239',
    'owner': '\u6d4b\u8bd5\u8239\u4e1c',
    'dwt': '25000DWT',
    'build_status': '\u5df2\u4ea4\u4ed8',
    'delivery_time': '2026\u5e74',
    'actual_delivery_date': '2026-06-18',
    'status_note': '\u51fa\u5382\u6295\u8fd0',
}
newbuilding_resp = client.post('/api/market-skill/newbuilding', json={'kind': 'newbuilding', 'record': newbuilding_record}, headers=auth_headers)
assert newbuilding_resp.status_code == 200
newbuilding_id = newbuilding_resp.json()['record']['id']
try:
    newbuilding_report_resp = client.post(
        '/api/market-skill/report',
        json={
            'kind': 'newbuilding',
            'period': 'weekly',
            'start_date': '2026-06-01',
            'end_date': '2026-06-30',
            'filters': {'q': 'smoke-newbuilding-1'},
        },
        headers=auth_headers,
    )
    assert newbuilding_report_resp.status_code == 200
    newbuilding_report = newbuilding_report_resp.json()
    assert newbuilding_report['source_stats']['delivered'] == 1
    assert newbuilding_report['detail_rows'][0]['ship_name'] == 'smoke-newbuilding-1'
    newbuilding_export_resp = client.post(
        '/api/export/market-skill-report',
        json={'kind': 'newbuilding', 'report': newbuilding_report, 'format': 'xlsx'},
        headers=auth_headers,
    )
    assert newbuilding_export_resp.status_code == 200
    wb7 = load_workbook(BytesIO(newbuilding_export_resp.content))
    assert wb7['分析报告']['A1'].value == '新造船市场分析报告'
finally:
    cleanup_market('newbuilding', newbuilding_id)

bad_docx_resp = client.post(
    '/api/parse',
    files={
        'file': (
            'bad.docx',
            b'not-a-real-docx',
            'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
        )
    },
    headers=auth_headers,
)
assert bad_docx_resp.status_code == 400
assert 'DOCX 文件格式无效' in bad_docx_resp.json()['detail']

original_config = load_config()
feishu_cargo_id = ''
feishu_newbuilding_id = ''
feishu_query_cargo_id = ''
feishu_query_newbuilding_id = ''
feishu_project_id = ''
feishu_query_project_id = ''
try:
    config_resp = client.post(
        '/api/config',
        json={
            **original_config,
            'base_url': '',
            'api_key': '',
            'model': '',
            'feishu_enabled': True,
            'feishu_receive_mode': 'http',
            'feishu_app_id': 'cli_smoke',
            'feishu_app_secret': 'secret-smoke',
            'feishu_verification_token': 'verify-smoke',
            'feishu_encrypt_key': 'encrypt-smoke',
        },
        headers=auth_headers,
    )
    assert config_resp.status_code == 200
    masked_config = config_resp.json()
    assert masked_config['has_feishu_app_secret'] is True
    assert masked_config['feishu_app_secret'] == '********'

    verify_resp = client.post(
        '/api/feishu/events',
        json={'type': 'url_verification', 'token': 'verify-smoke', 'challenge': 'challenge-ok'},
    )
    assert verify_resp.status_code == 200
    assert verify_resp.json()['challenge'] == 'challenge-ok'

    unavailable_reply = handle_feishu_message({'open_id': 'ou_anyone', 'chat_id': 'chat_anywhere', 'message_id': 'mid-help', 'text': '你好', 'files': []}, load_config())
    assert 'AI 暂不可用' in unavailable_reply
    assert '我现在能做这些事' not in unavailable_reply
    ws_event = type(
        'WsEvent',
        (),
        {
            'header': type('Header', (), {'event_id': 'ws-event', 'event_type': 'im.message.receive_v1', 'token': ''})(),
            'event': type(
                'Event',
                (),
                {
                    'sender': type('Sender', (), {'sender_id': type('SenderId', (), {'open_id': 'ou_ws', 'user_id': '', 'union_id': ''})()})(),
                    'message': type('Message', (), {'message_id': 'mid-ws', 'chat_id': 'chat_ws', 'message_type': 'text', 'content': '{"text":"帮助"}'})(),
                },
            )(),
        },
    )()
    ws_payload = ws_message_to_payload(ws_event)
    assert ws_payload['header']['event_type'] == 'im.message.receive_v1'
    assert ws_payload['event']['message']['content'] == '{"text":"帮助"}'

    save_config({**load_config(), 'base_url': 'https://example.com/v1', 'api_key': 'sk-agent-test', 'model': 'agent-test'})

    def fake_agent_chat(config, *, system_prompt, user_prompt, temperature=None):
        payload = json.loads(user_prompt)
        text = payload.get('user_text', '')
        if 'required_json' in payload:
            if '4000吨甲苯' in text:
                return json.dumps({'reply': '', 'tool_calls': [{'name': 'extract_cargo_opportunity', 'arguments': {'text': text}}], 'needs_confirmation': True, 'pending_update': {}}, ensure_ascii=False)
            if '3000 吨甲苯' in text:
                return json.dumps({'reply': '', 'tool_calls': [{'name': 'chat_general', 'arguments': {}}], 'needs_confirmation': False, 'pending_update': {}}, ensure_ascii=False)
            if '5000吨甲醇' in text:
                return json.dumps({'reply': '', 'tool_calls': [{'name': 'extract_cargo_opportunity', 'arguments': {'text': text}}], 'needs_confirmation': True, 'pending_update': {}}, ensure_ascii=False)
            if '6000吨乙二醇' in text:
                return json.dumps({'reply': '', 'tool_calls': [{'name': 'extract_cargo_opportunity', 'arguments': {'text': text}}], 'needs_confirmation': True, 'pending_update': {}}, ensure_ascii=False)
            if '烟测新船88' in text:
                return json.dumps({'reply': '', 'tool_calls': [{'name': 'extract_newbuilding_info', 'arguments': {'text': text}}], 'needs_confirmation': True, 'pending_update': {}}, ensure_ascii=False)
            if '只有烟测船厂' in text:
                return json.dumps({'reply': '', 'tool_calls': [{'name': 'extract_newbuilding_info', 'arguments': {'text': text}}], 'needs_confirmation': True, 'pending_update': {}}, ensure_ascii=False)
            if '装港改成' in text:
                return json.dumps({'reply': '', 'tool_calls': [{'name': 'update_pending_record', 'arguments': {'text': text}}], 'needs_confirmation': True, 'pending_update': {}}, ensure_ascii=False)
            if text in {'确认保存', '保存'}:
                return json.dumps({'reply': '', 'tool_calls': [{'name': 'save_pending_record', 'arguments': {}}], 'needs_confirmation': False, 'pending_update': {}}, ensure_ascii=False)
            if '保存了吗' in text:
                return json.dumps({'reply': '', 'tool_calls': [{'name': 'check_last_saved_record', 'arguments': {}}], 'needs_confirmation': False, 'pending_update': {}}, ensure_ascii=False)
            if text == '保存项目':
                return json.dumps({'reply': '', 'tool_calls': [{'name': 'save_bid_project', 'arguments': {}}], 'needs_confirmation': False, 'pending_update': {}}, ensure_ascii=False)
            if '解析标书' in text:
                return json.dumps({'reply': '', 'tool_calls': [{'name': 'parse_bid_file', 'arguments': {}}], 'needs_confirmation': False, 'pending_update': {}}, ensure_ascii=False)
            if '截止时间是什么' in text:
                return json.dumps({'reply': '', 'tool_calls': [{'name': 'answer_bid_question', 'arguments': {'question': text}}], 'needs_confirmation': False, 'pending_update': {}}, ensure_ascii=False)
            if '最近张家港到东莞甲苯货盘' in text:
                return json.dumps({'reply': '', 'tool_calls': [{'name': 'search_market_records', 'arguments': {'kind': 'cargo', 'query': '甲苯'}}], 'needs_confirmation': False, 'pending_update': {}}, ensure_ascii=False)
            if '跟进中的甲苯货盘' in text:
                return json.dumps({'reply': '', 'tool_calls': [{'name': 'search_market_records', 'arguments': {'kind': 'cargo', 'query': '甲苯', 'status': '跟进中'}}], 'needs_confirmation': False, 'pending_update': {}}, ensure_ascii=False)
            if '烟测船厂新造船' in text:
                return json.dumps({'reply': '', 'tool_calls': [{'name': 'search_market_records', 'arguments': {'kind': 'newbuilding', 'query': '烟测船厂'}}], 'needs_confirmation': False, 'pending_update': {}}, ensure_ascii=False)
            if '完全不存在的烟测货盘' in text:
                return json.dumps({'reply': '', 'tool_calls': [{'name': 'search_market_records', 'arguments': {'kind': 'cargo', 'query': '完全不存在的烟测货盘'}}], 'needs_confirmation': False, 'pending_update': {}}, ensure_ascii=False)
            if '福海创以前有没有项目' in text:
                return json.dumps({'reply': '', 'tool_calls': [{'name': 'search_project', 'arguments': {'query': '福海创'}}], 'needs_confirmation': False, 'pending_update': {}}, ensure_ascii=False)
            return json.dumps({'reply': '', 'tool_calls': [{'name': 'chat_general', 'arguments': {}}], 'needs_confirmation': False, 'pending_update': {}}, ensure_ascii=False)
        tool_results = payload.get('tool_results') or []
        if tool_results and tool_results[0].get('tool') == 'extract_cargo_opportunity':
            if tool_results[0].get('missing'):
                return tool_results[0].get('reply') or '请补充缺失字段。'
            record = tool_results[0].get('record') or {}
            return f"我已帮你填好商机草稿，装港：{record.get('load_port') or '-'}，卸港：{record.get('discharge_port') or '-'}，请确认保存。"
        if tool_results and tool_results[0].get('tool') == 'extract_newbuilding_info':
            if tool_results[0].get('missing'):
                return tool_results[0].get('reply') or '请补充缺失字段。'
            record = tool_results[0].get('record') or {}
            return f"我已帮你填好新造船草稿，船名：{record.get('ship_name') or '-'}，船厂：{record.get('shipyard') or '-'}，请确认保存。"
        if tool_results and tool_results[0].get('tool') == 'update_pending_record':
            return f"已更新草稿，装港：{tool_results[0]['record'].get('load_port')}"
        if tool_results and tool_results[0].get('tool') == 'save_pending_record':
            if not tool_results[0].get('ok'):
                return tool_results[0].get('error') or '保存失败'
            return f"已保存，记录 ID：{tool_results[0]['record']['id']}"
        if tool_results and tool_results[0].get('tool') == 'check_last_saved_record':
            return tool_results[0].get('reply') or '未找到保存状态'
        if tool_results and tool_results[0].get('tool') == 'save_bid_project':
            if not tool_results[0].get('ok'):
                return tool_results[0].get('error') or '项目保存失败'
            return f"项目已保存，项目 ID：{tool_results[0]['saved_project_id']}"
        if tool_results and tool_results[0].get('tool') == 'answer_bid_question':
            return tool_results[0].get('reply') or '未找到标书答案'
        if tool_results and tool_results[0].get('tool') == 'parse_bid_file':
            if not tool_results[0].get('ok'):
                return tool_results[0].get('error') or '标书解析失败'
            return tool_results[0].get('reply') or '标书解析完成'
        if tool_results and tool_results[0].get('tool') == 'search_market_records':
            return tool_results[0].get('reply') or '未查到市场情报'
        if tool_results and tool_results[0].get('tool') == 'search_project':
            return tool_results[0].get('reply') or '未查到历史项目'
        return '这是 AI 自然回复，不是固定菜单。'

    import app as app_module
    original_module_call_chat = app_module.call_chat_completion
    app_module.call_chat_completion = fake_agent_chat
    try:
        chat_reply = handle_feishu_message({'open_id': 'ou_chat', 'chat_id': 'chat_agent_chat', 'message_id': 'mid-chat', 'text': '你好，随便聊聊', 'files': []}, load_config())
        assert chat_reply == '这是 AI 自然回复，不是固定菜单。'
        assert '我现在能做这些事' not in chat_reply
        save_config({**load_config(), 'feishu_allowed_open_ids': 'ou_allowed', 'feishu_allowed_chat_ids': ''})
        unauthorized_reply = handle_feishu_message({'open_id': 'ou_blocked', 'chat_id': 'chat_blocked', 'message_id': 'mid-blocked', 'text': '你好', 'files': []}, load_config())
        assert '没有权限' in unauthorized_reply
        save_config({**load_config(), 'feishu_allowed_open_ids': '', 'feishu_allowed_chat_ids': ''})

        message = {
            'open_id': 'ou_smoke',
            'chat_id': 'chat_smoke',
            'message_id': 'mid-smoke',
            'text': '4000吨甲苯，张家港～东莞，要求月内装出，远大的计划',
            'files': [],
        }
        reply = handle_feishu_message(message, load_config())
        assert '商机草稿' in reply
        fallback_reply = handle_feishu_message({**message, 'open_id': 'ou_fallback', 'chat_id': 'chat_fallback', 'text': '3000 吨甲苯，张家港到东莞，月底装'}, load_config())
        assert '商机草稿' in fallback_reply
        assert '张家港' in fallback_reply
        assert '东莞' in fallback_reply
        clarify_reply = handle_feishu_message({**message, 'open_id': 'ou_clarify', 'chat_id': 'chat_clarify', 'text': '5000吨甲醇，宁波装，月底装'}, load_config())
        assert '请补充' in clarify_reply
        assert '卸港' in clarify_reply
        phrase_reply = handle_feishu_message({**message, 'open_id': 'ou_phrase', 'chat_id': 'chat_phrase', 'text': '6000吨乙二醇，宁波装，东莞卸，月底装，货主远大'}, load_config())
        assert '商机草稿' in phrase_reply
        assert '宁波' in phrase_reply
        assert '东莞' in phrase_reply
        update_reply = handle_feishu_message({**message, 'text': '装港改成宁波'}, load_config())
        assert '宁波' in update_reply
        confirm_reply = handle_feishu_message({**message, 'text': '确认保存'}, load_config())
        feishu_cargo_id = re.search(r'[a-f0-9]{32}', confirm_reply).group(0)
        assert '已保存' in confirm_reply
        saved_record_path = market_record_path('cargo', feishu_cargo_id)
        assert saved_record_path.exists()
        saved_record = json.loads(saved_record_path.read_text(encoding='utf-8'))
        assert saved_record['source'] == '飞书'
        assert saved_record['load_port'] == '宁波'
        assert saved_record['source_channel'] == 'feishu'
        assert saved_record['source_message_id'] == 'mid-smoke'
        assert saved_record['confirmed_at']
        list_saved_resp = client.get('/api/market-skill/cargo', params={'q': '甲苯'}, headers=auth_headers)
        assert list_saved_resp.status_code == 200
        listed_saved = next(item for item in list_saved_resp.json()['items'] if item['id'] == feishu_cargo_id)
        assert listed_saved['source'] == '飞书'
        assert listed_saved['source_channel'] == 'feishu'
        assert '4000吨甲苯' in listed_saved['source_text']
        status_reply = handle_feishu_message({**message, 'text': '刚才那条保存了吗'}, load_config())
        assert feishu_cargo_id in status_reply
        assert '已经保存' in status_reply

        newbuilding_message = {
            'open_id': 'ou_newbuilding_save',
            'chat_id': 'chat_newbuilding_save',
            'message_id': 'mid-newbuilding-save',
            'text': '烟测新船88，烟测船厂为烟测船东建造12000DWT化学品船，预计2027年交付',
            'files': [],
        }
        newbuilding_reply = handle_feishu_message(newbuilding_message, load_config())
        assert '新造船草稿' in newbuilding_reply
        assert '烟测新船88' in newbuilding_reply
        newbuilding_clarify_reply = handle_feishu_message({**newbuilding_message, 'open_id': 'ou_newbuilding_clarify', 'chat_id': 'chat_newbuilding_clarify', 'text': '只有烟测船厂有新造船消息'}, load_config())
        assert '请补充' in newbuilding_clarify_reply
        assert '船名' in newbuilding_clarify_reply
        newbuilding_confirm_reply = handle_feishu_message({**newbuilding_message, 'text': '确认保存'}, load_config())
        feishu_newbuilding_id = re.search(r'[a-f0-9]{32}', newbuilding_confirm_reply).group(0)
        saved_newbuilding_path = market_record_path('newbuilding', feishu_newbuilding_id)
        assert saved_newbuilding_path.exists()
        saved_newbuilding = json.loads(saved_newbuilding_path.read_text(encoding='utf-8'))
        assert saved_newbuilding['ship_name'] == '烟测新船88'
        assert saved_newbuilding['source'] == '飞书'
        assert saved_newbuilding['source_channel'] == 'feishu'
        assert saved_newbuilding['source_message_id'] == 'mid-newbuilding-save'
        list_newbuilding_resp = client.get('/api/market-skill/newbuilding', params={'q': '烟测新船88'}, headers=auth_headers)
        assert list_newbuilding_resp.status_code == 200
        listed_newbuilding = next(item for item in list_newbuilding_resp.json()['items'] if item['id'] == feishu_newbuilding_id)
        assert listed_newbuilding['source_channel'] == 'feishu'
        assert '烟测新船88' in listed_newbuilding['source_text']

        original_download_file = app_module.feishu_download_file
        original_parse_ai_document = app_module.parse_ai_document

        def fake_download_file(config, message_id, file_key, file_name=''):
            with tempfile.NamedTemporaryFile(delete=False, suffix='.docx') as temp_file:
                temp_path = Path(temp_file.name)
            document_text = '飞书文件解析项目 招标编号 FS-PARSE-001 投标截止时间 2026-09-01 09:30 保证金 10000 元。' * 6
            xml = (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
                '<w:body><w:p><w:r><w:t>'
                + document_text
                + '</w:t></w:r></w:p></w:body></w:document>'
            )
            with ZipFile(temp_path, 'w') as archive:
                archive.writestr('word/document.xml', xml)
            return temp_path

        def fake_parse_ai_document(document_text, config):
            assert '飞书文件解析项目' in document_text
            return {
                'document_summary': {
                    'project_name': '飞书文件解析项目',
                    'bid_no': 'FS-PARSE-001',
                    'tenderer': '飞书招标人',
                    'bid_deadline': '2026-09-01 09:30',
                    'deposit_amount': '10000 元',
                    'qualification_requirements': '具备烟测资格',
                },
                'extraction_fields': {},
                'register_rows': [],
                'analysis': {'summary': '飞书文件解析摘要'},
                'match_review': [{'item': '资格', 'status': '注意', 'reason': '烟测风险', 'action': '复核'}],
            }

        app_module.feishu_download_file = fake_download_file
        app_module.parse_ai_document = fake_parse_ai_document
        try:
            bid_parse_reply = handle_feishu_message(
                {
                    'open_id': 'ou_bid_parse',
                    'chat_id': 'chat_bid_parse',
                    'message_id': 'mid-bid-parse',
                    'text': '解析标书',
                    'files': [{'file_key': 'file-smoke', 'file_name': 'feishu-smoke.docx'}],
                },
                load_config(),
            )
            assert '标书解析完成' in bid_parse_reply
            assert '飞书文件解析项目' in bid_parse_reply
            parsed_session = app_module.feishu_dialog_load('ou_bid_parse', 'chat_bid_parse')
            assert parsed_session['active_skill'] == 'bid_parse'
            assert parsed_session['last_result']['document_summary']['project_name'] == '飞书文件解析项目'
            assert parsed_session['last_file_name'] == 'feishu-smoke.docx'
            bad_parse_reply = handle_feishu_message(
                {
                    'open_id': 'ou_bid_bad_parse',
                    'chat_id': 'chat_bid_bad_parse',
                    'message_id': 'mid-bid-bad-parse',
                    'text': '解析标书',
                    'files': [{'file_key': 'file-bad', 'file_name': 'bad.exe'}],
                },
                load_config(),
            )
            assert '当前只支持' in bad_parse_reply
        finally:
            app_module.feishu_download_file = original_download_file
            app_module.parse_ai_document = original_parse_ai_document
        recent_logs = sorted(FEISHU_AGENT_LOGS_DIR.glob('*/*.json'), key=lambda path: path.stat().st_mtime, reverse=True)[:10]
        assert recent_logs
        log_text = '\n'.join(path.read_text(encoding='utf-8') for path in recent_logs)
        assert 'mid-smoke' in log_text
        assert 'sk-agent-test' not in log_text
        logs_resp = client.get('/api/feishu/agent-logs?limit=10', headers=auth_headers)
        assert logs_resp.status_code == 200
        assert logs_resp.json()['count'] > 0
        assert 'sk-agent-test' not in json.dumps(logs_resp.json(), ensure_ascii=False)
        logs_export_resp = client.get('/api/export/feishu-agent-logs?limit=10', headers=auth_headers)
        assert logs_export_resp.status_code == 200
        assert b'mid-smoke' in logs_export_resp.content

        import app as app_module_for_bid
        bid_session = app_module_for_bid.feishu_dialog_load('ou_bid_save', 'chat_bid_save')
        bid_session.update(
            {
                'active_skill': 'bid_parse',
                'last_result': {
                    'document_summary': {
                        'project_name': '飞书保存烟测项目',
                        'bid_no': 'FS-SMOKE-001',
                        'tenderer': '烟测招标人',
                        'bid_deadline': '2026-08-01 10:00',
                    },
                    'extraction_fields': {},
                    'register_rows': [],
                    'analysis': {'summary': '烟测解析摘要'},
                    'match_review': [],
                },
                'last_file_name': 'feishu-smoke.pdf',
                'last_parse_text': '解析标书',
                'last_question': '解析标书',
            }
        )
        app_module_for_bid.feishu_dialog_save('ou_bid_save', 'chat_bid_save', bid_session)
        bid_question_reply = handle_feishu_message({'open_id': 'ou_bid_save', 'chat_id': 'chat_bid_save', 'message_id': 'mid-bid-question', 'text': '截止时间是什么', 'files': []}, load_config())
        assert '2026-08-01 10:00' in bid_question_reply
        bid_save_reply = handle_feishu_message({'open_id': 'ou_bid_save', 'chat_id': 'chat_bid_save', 'message_id': 'mid-bid-save', 'text': '保存项目', 'files': []}, load_config())
        feishu_project_id = re.search(r'[a-f0-9]{32}', bid_save_reply).group(0)
        assert '项目已保存' in bid_save_reply
        assert project_path(feishu_project_id).exists()
        saved_project_resp = client.get(f'/api/projects/{feishu_project_id}', headers=auth_headers)
        assert saved_project_resp.status_code == 200
        saved_project = saved_project_resp.json()
        assert saved_project['result']['document_summary']['project_name'] == '飞书保存烟测项目'
        assert saved_project['follow_up']['information_source'] == '飞书'
        assert saved_project['source_channel'] == 'feishu'
        assert saved_project['source_open_id'] == 'ou_bid_save'
        assert saved_project['source_chat_id'] == 'chat_bid_save'
        assert saved_project['source_message_id'] == 'mid-bid-save'
        assert saved_project['source_text'] == '解析标书'
        assert saved_project['confirmed_at']

        query_record_resp = client.post(
            '/api/market-skill/cargo',
            json={
                'kind': 'cargo',
                'record': {
                    'board_type': '即时货盘',
                    'segment': '内贸化',
                    'cargo_name': '甲苯',
                    'tonnage': '3000吨',
                    'load_port': '张家港',
                    'discharge_port': '东莞',
                    'cargo_owner': '烟测货主',
                    'status': '跟进中',
                },
            },
            headers=auth_headers,
        )
        assert query_record_resp.status_code == 200
        feishu_query_cargo_id = query_record_resp.json()['record']['id']
        market_query_reply = handle_feishu_message({'open_id': 'ou_market_query', 'chat_id': 'chat_market_query', 'message_id': 'mid-market-query', 'text': '最近张家港到东莞甲苯货盘有哪些', 'files': []}, load_config())
        assert feishu_query_cargo_id in market_query_reply
        assert '甲苯' in market_query_reply
        market_status_query_reply = handle_feishu_message({'open_id': 'ou_market_status_query', 'chat_id': 'chat_market_status_query', 'message_id': 'mid-market-status-query', 'text': '跟进中的甲苯货盘有哪些', 'files': []}, load_config())
        assert feishu_query_cargo_id in market_status_query_reply
        assert '跟进中' in market_status_query_reply

        newbuilding_query_resp = client.post(
            '/api/market-skill/newbuilding',
            json={
                'kind': 'newbuilding',
                'record': {
                    'stage': '预计2027年完造出厂',
                    'ship_name': '烟测新船01',
                    'shipyard': '烟测船厂',
                    'owner': '烟测船东',
                    'dwt': '12000DWT',
                    'ship_type': '化学品船',
                },
            },
            headers=auth_headers,
        )
        assert newbuilding_query_resp.status_code == 200
        feishu_query_newbuilding_id = newbuilding_query_resp.json()['record']['id']
        newbuilding_query_reply = handle_feishu_message({'open_id': 'ou_newbuilding_query', 'chat_id': 'chat_newbuilding_query', 'message_id': 'mid-newbuilding-query', 'text': '烟测船厂新造船有哪些', 'files': []}, load_config())
        assert feishu_query_newbuilding_id in newbuilding_query_reply
        assert '烟测船厂' in newbuilding_query_reply

        no_result_reply = handle_feishu_message({'open_id': 'ou_market_empty', 'chat_id': 'chat_market_empty', 'message_id': 'mid-market-empty', 'text': '完全不存在的烟测货盘有哪些', 'files': []}, load_config())
        assert '没有查到' in no_result_reply
        assert '换个' in no_result_reply

        feishu_query_project_id = uuid.uuid4().hex
        save_project(
            feishu_query_project_id,
            ProjectPayload(
                title='福海创烟测历史项目',
                source_file_name='fuhai-smoke.pdf',
                result={'document_summary': {'project_name': '福海创烟测历史项目', 'bid_no': 'FHC-SMOKE-001', 'tenderer': '福海创'}},
                follow_up={'bid_status': '已投标', 'award_status': '未知'},
            ),
        )
        project_query_reply = handle_feishu_message({'open_id': 'ou_project_query', 'chat_id': 'chat_project_query', 'message_id': 'mid-project-query', 'text': '福海创以前有没有项目', 'files': []}, load_config())
        assert '福海创烟测历史项目' in project_query_reply
        assert 'FHC-SMOKE-001' in project_query_reply
    finally:
        app_module.call_chat_completion = original_module_call_chat
finally:
    if feishu_cargo_id:
        cleanup_market('cargo', feishu_cargo_id)
    if feishu_newbuilding_id:
        cleanup_market('newbuilding', feishu_newbuilding_id)
    if feishu_query_cargo_id:
        cleanup_market('cargo', feishu_query_cargo_id)
    if feishu_query_newbuilding_id:
        cleanup_market('newbuilding', feishu_query_newbuilding_id)
    if feishu_project_id:
        cleanup(feishu_project_id)
    if feishu_query_project_id:
        cleanup(feishu_query_project_id)
    save_config(original_config)

users_resp = client.get('/api/users', headers=auth_headers)
assert users_resp.status_code == 200
assert any(item['username'] == 'ruico' for item in users_resp.json()['items'])

create_user_resp = client.post(
    '/api/users',
    json={'username': 'demo_user', 'password': 'Demo123@', 'display_name': '演示用户', 'role': 'user'},
    headers=auth_headers,
)
assert create_user_resp.status_code == 200

delete_resp = client.delete(f'/api/projects/{created_id}', headers=auth_headers)
assert delete_resp.status_code == 200
cleanup(created_id)
cleanup_user('demo_user')
